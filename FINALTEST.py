from flask import Flask, render_template, request, jsonify, send_file
import os
import pdfplumber
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)

# Function to extract text from PDF
def extract_text_from_pdf(pdf_path):
    text_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                lines = text.split('\n')
                text_lines.extend(lines)
    return text_lines

# Function to extract information from text
def extract_info_from_text(text_lines):
    info = {
        "Program Section:": "",
        "Subject Code & Title:": "",
        "Test Name:": ""
    }
    for line in text_lines:
        for key in info.keys():
            if line.startswith(key):
                info[key] = line
    return info

# Function to extract marks from text
def extract_marks_from_text(text_lines, reg_numbers):
    marks_dict = {}
    for line in text_lines:
        for reg_number in reg_numbers:
            if reg_number in line:
                parts = line.split()
                reg_index = parts.index(reg_number)
                if reg_index + 1 < len(parts):
                    marks_dict[reg_number] = parts[reg_index + 1]
    return marks_dict

# Function to process PDFs in a folder
def process_pdfs_in_folder(file_order, folder_path, reg_numbers):
    results = {}
    info = {
        "Program Section:": "",
        "Subject Code & Title:": "",
        "Test Name:": ""
    }
    for filename in file_order:
        pdf_path = os.path.join(folder_path, filename)
        text_lines = extract_text_from_pdf(pdf_path)
        marks = extract_marks_from_text(text_lines, reg_numbers)
        for reg_number, marks_value in marks.items():
            if reg_number not in results:
                results[reg_number] = {}
            results[reg_number][filename] = marks_value
        extracted_info = extract_info_from_text(text_lines)
        for key in info:
            if extracted_info[key]:
                info[key] = extracted_info[key]
    return results, info

# Function to create DataFrame for component split-up
def input_split_up(components, array_2d):
    split_up_df = pd.DataFrame(columns=['Component', 'CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6'])
    split_up_df['Component'] = components

    for index, row in split_up_df.iterrows():
        for i, co in enumerate(['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']):
            split_up_df.at[index, co] = array_2d[index][i]

    return split_up_df

# Function to calculate total possible attainment
def calculate_total_possible_attainment(split_up_df):
    total_attainment = split_up_df[['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']].sum()
    return total_attainment

# Function to calculate row-wise sum
def calculate_row_wise_sum(split_up_df):
    split_up_df['Row-wise Sum'] = split_up_df.iloc[:, 1:].sum(axis=1)
    return split_up_df

# Function to calculate attainment
def calculate_attainment(split_up_df, marks_df, reg_numbers):
    attainment_df = pd.DataFrame(columns=['Register number', 'CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6'])
    attainment_df['Register number'] = reg_numbers

    for index, reg_number in enumerate(reg_numbers):
        attainment_values = {co: 0.0 for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']}
        for component in marks_df.columns:
            if component in split_up_df['Component'].values:
                student_marks = marks_df.loc[reg_number, component]
                row_sum = split_up_df.loc[split_up_df['Component'] == component, 'Row-wise Sum'].values[0]
                try:
                    student_marks = float(student_marks)
                    percentage_scored = (student_marks / row_sum) * 100
                except ValueError:
                    percentage_scored = 0

                for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']:
                    split_value = split_up_df.loc[split_up_df['Component'] == component, co].values[0]
                    attainment_values[co] += (percentage_scored * split_value) / 100

        for co in ['CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6']:
            attainment_df.at[index, co] = attainment_values[co]

    return attainment_df
def calculate_students_above_target(attainment_df, target_percentage, total_attainment):
    # Calculate the threshold based on the target percentage
    threshold_values = {co: (target_percentage / 100) * total_attainment[co] for co in total_attainment.index}

    # Prepare a DataFrame to store results
    above_target_df = pd.DataFrame(columns=['CO', 'No. of Students', 'Percentage_of_Students'])

    total_students = len(attainment_df)

    for co in attainment_df.columns[1:]:  # Skip the 'Register number' column
        threshold = threshold_values[co]
        # Calculate the number of students who scored above the threshold
        num_students_above = len(attainment_df[attainment_df[co] >= threshold])
        # Calculate the percentage of students who scored above the threshold
        percentage_above = (num_students_above / total_students) * 100
        # Append the results to the DataFrame
        above_target_df = pd.concat([above_target_df, pd.DataFrame([[co, num_students_above, percentage_above]], columns=['CO', 'No. of Students', 'Percentage_of_Students'])])

    return above_target_df

def get_attainment_levels(dict):
    levels = {}
    for level in [3, 2, 1, 0]:
        while True:
            try:
                interval = dict[str(level)].split('-')
                low, high = float(interval[0]), float(interval[1])
                levels[level] = (low, high)
                break
            except ValueError:
                print("Please enter a valid percentage range.")

    return levels

def assign_attainment_levels(above_target_df, levels):
    # Print the columns to verify correct naming
    print(above_target_df.columns)

    # Assign levels based on intervals
    def assign_level(percentage):
        for level in sorted(levels.keys(), reverse=True):
            if levels[level][0] <= percentage <= levels[level][1]:
                return level
        return 0

    # Apply the level assignment for each CO based on the percentage of students above the target
    if 'Percentage_of_Students' in above_target_df.columns:
        above_target_df['Attainment Level'] = above_target_df['Percentage_of_Students'].apply(assign_level)
    else:
        raise KeyError("Column 'Percentage_of_Students' not found in above_target_df")

    # Return the Attainment Level column as a list (single-line array)
    return above_target_df['Attainment Level'].tolist()

def get_clo_pso_input():
    clo_pso_data = []
    for i in range(1, 7):  # CLO1 to CLO6
        clo_row = [f'CLO{i}']
        for j in range(1, 13):  # PO1 to PO12
            value = float(input(f"Enter value for CLO{i}, PO{j}: "))
            clo_row.append(value)
        for j in range(1, 4):  # PSO1 to PSO3
            value = float(input(f"Enter value for CLO{i}, PSO{j}: "))
            clo_row.append(value)
        clo_pso_data.append(clo_row)

    # Create DataFrame from input
    columns = ['CLO'] + [f'PO{i}' for i in range(1, 13)] + [f'PSO{i}' for i in range(1, 4)]
    df_clo_pso = pd.DataFrame(clo_pso_data, columns=columns)

    return df_clo_pso

def calculate_targets_and_attained(df_clo_pso, attainment_levels_df):
    # Ensure the DataFrame is numeric and handle missing values
    df_clo_pso = df_clo_pso.apply(pd.to_numeric, errors='coerce')
    attainment_levels_df = [float(x) for x in attainment_levels_df]
    
    # Initialize lists for targets and attained values
    targets = df_clo_pso.mean()
    attained_values = []

    # For each PO and PSO
    for column in df_clo_pso.columns:
        total_sum = 0
        count_non_nan = 0  # Count of non-NaN values for the current column
        
        for i in range(len(df_clo_pso)):
            value = df_clo_pso.iloc[i][column]
            if not pd.isna(value):  # Only consider non-NaN values
                attainment = attainment_levels_df[i]
                total_sum += (value * attainment) / 3
                count_non_nan += 1
        
        # Avoid division by zero if all values are NaN
        if count_non_nan > 0:
            attained_value = total_sum / count_non_nan
        else:
            attained_value = np.nan  # Or some other placeholder for no valid values
        
        attained_values.append(attained_value)

    # Create DataFrames for targets and attained values
    df_attained = pd.DataFrame([attained_values], columns=df_clo_pso.columns, index=['Attained'])
    df_final = pd.DataFrame(columns=df_attained.columns)
    
    # Handle NaN values in targets
    if targets.isnull().any():
        print("Warning: Target values contain NaNs. They will be replaced with 0.")
        targets = targets.fillna(0)
    
    df_final.loc['Target'] = targets
    df_final.loc['Attained'] = df_attained.loc['Attained']

    # Debug: Print intermediate values
    print("Targets:", targets)
    print("Attained Values:", attained_values)
    print(df_attained)
    print(df_final)

    return targets, df_final

def save_to_excel(marks_df, info, total_attainment, attainment_df, split_up_df_without_sum, df_clo_pso, targets, df_attained, above_target_df, attainment_levels_df, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write the marks DataFrame to the first sheet
        marks_df.to_excel(writer, sheet_name='Extracted Marks')

        # Create the 'Attainment and Analysis' sheet
        sheet_name = 'Attainment and Analysis'
        workbook = writer.book
        writer.sheets[sheet_name] = workbook.create_sheet(sheet_name)

        # Define the starting row positions for each DataFrame
        start_row_register = 0
        start_row_split_up = start_row_register + len(marks_df) + 5  # 5 rows gap after the 'Register number' table
        start_row_total_attainment = start_row_split_up + len(split_up_df_without_sum)
        start_row_attainment = start_row_total_attainment + 5  # 5 rows gap after average attainment row
        start_row_clo_pso = start_row_attainment + len(attainment_df) + 6  # 6 rows gap after attainment_df
        start_row_attained = start_row_clo_pso + len(df_clo_pso) + 2  # 2 rows gap after df_clo_pso

        # Write the split_up_df_without_sum DataFrame
        split_up_df_without_sum.to_excel(writer, sheet_name=sheet_name, startrow=start_row_split_up, index=False)
        
        # Load the workbook to append the total possible attainment row
        sheet = workbook[sheet_name]
        
        # Identify the end row of the "Register number" table
        register_number_end_row = start_row_split_up + len(marks_df) + 1  # Adjust based on actual rows

        # Write the "Total Possible Attainment" row
        total_row = ['Total Possible Attainment'] + total_attainment.tolist()
        for col_num, value in enumerate(total_row, 1):  # start from column 1 (A)
            sheet.cell(row=start_row_total_attainment + 1, column=col_num).value = value

        # Write the "No. of Students Above Target" and "Percentage of Students Above Target" rows
        above_target_row1 = ['No. of Students Above Target'] + above_target_df['No. of Students'].tolist()
        above_target_row2 = ['Percentage of Students Above Target'] + above_target_df['Percentage_of_Students'].tolist()
        
        for col_num, value in enumerate(above_target_row1, 1):  # start from column 1 (A)
            sheet.cell(row=start_row_total_attainment + 2, column=col_num).value = value
        for col_num, value in enumerate(above_target_row2, 1):  # start from column 1 (A)
            sheet.cell(row=start_row_total_attainment + 3, column=col_num).value = value

        # Write the "Attainment Levels" row directly under "Percentage of Students Above Target"
        if isinstance(attainment_levels_df, list):
            attainment_levels_df = ['Attainment Levels'] + attainment_levels_df

        for col_num, value in enumerate(attainment_levels_df, 1):  # start from column 1 (A)
            sheet.cell(row=start_row_total_attainment + 4, column=col_num).value = value

        # Calculate the average of the six attainment values
        attainment_values = attainment_levels_df[1:]  # skip the label
        average_attainment = sum(attainment_values) / len(attainment_values)

        # Write the "Average Attainment" row directly under "Attainment Levels"
        sheet.cell(row=start_row_total_attainment + 5, column=1).value = 'Average Attainment'
        sheet.cell(row=start_row_total_attainment + 5, column=2).value = average_attainment

        # Write the other DataFrames to the specified rows in the sheet
        attainment_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row_attainment, index=False)
        df_clo_pso.to_excel(writer, sheet_name=sheet_name, startrow=start_row_clo_pso, index=False)
        df_attained.to_excel(writer, sheet_name=sheet_name, startrow=start_row_attained)

    # After writing the DataFrames, use openpyxl to format the sheet
    workbook = load_workbook(output_path)
    sheet = workbook[sheet_name]
    
    
    marks_sheet = workbook['Extracted Marks']
    marks_sheet.column_dimensions['A'].width = 20
    for col in range(2, 10):  # Columns B to I
        marks_sheet.column_dimensions[chr(64 + col)].width = 15

    # Set column widths
    sheet.column_dimensions['A'].width = 20
    for col in range(2, 12):  # B to K
        sheet.column_dimensions[chr(64 + col)].width = 15

    # Insert and merge rows, and add header information
    sheet.insert_rows(1, 6)
    sheet.merge_cells('H1:M1')
    sheet['H1'] = 'SRM Institute of Science and Technology Vadapalani'
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells('H2:M2')
    sheet['H2'] = info.get('Program Section:', '')
    sheet.row_dimensions[2].height = 30

    sheet.merge_cells('H3:M3')
    sheet['H3'] = info.get('Subject Code & Title:', '')
    sheet.row_dimensions[3].height = 30

    # Save the changes to the workbook
    workbook.save(output_path)




@app.route('/')
def index():
    return render_template('index 3.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    registration_numbers = request.form.get('regNumbers')
    uploaded_files = request.files.getlist('pdfFiles')

    if not registration_numbers:
        return jsonify({'success': False, 'message': 'No registration numbers provided.'})

    if not uploaded_files:
        return jsonify({'success': False, 'message': 'No files uploaded.'})

    response = {
        'success': True,
        'message': 'Files uploaded successfully.'
    }

    return jsonify(response)

@app.route('/download/<filename>')
def download_file(filename):
    results_dir = './results'
    file_path = os.path.join(results_dir, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404

@app.route('/generate_cam', methods=['POST'])
def generate_cam():
    data = request.get_json()
    file_names = data.get('fileNames', [])

    cam_table = []
    for pdf_name in file_names:
        row = [pdf_name[:-4]] + ['' for _ in range(7)]  # Remove .pdf extension
        cam_table.append(row)

    return jsonify({'cam_table': cam_table}) 

@app.route('/calculateAttainment', methods=['POST'])
def calculate_attainment_route():
    try:
        reg_numbers = request.form.get('regNumbers')
        components_array = request.form.get('ComponentsArray')
        targetPercentage=request.form.get('targetPercentage')
        target_range=request.form.get('Target_range')
        COPOMapperTablevalues=request.form.get('COPOMapperTablevalues')
        
        file_order = request.form.get('fileOrder')
        if not reg_numbers or not components_array:
            return jsonify({'error': 'Missing regNumbers or ComponentsArray'})

        reg_numbers = json.loads(reg_numbers)
        components_array = json.loads(components_array)
        targetPercentage=json.loads(targetPercentage)
        target_range=json.loads(target_range)
        COPOMapperTablevalues=json.loads(COPOMapperTablevalues)
        targetPercentage=int(targetPercentage)
        print(COPOMapperTablevalues)
        print(target_range)
        print(type(targetPercentage))
        file_order = json.loads(file_order)
        pdf_files = request.files.getlist('pdfFiles')
        uploaded_files = pdf_files

        upload_dir = './uploads'
        results_dir = './results'
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)

        file_info = []
        for file in uploaded_files:
            file_path = os.path.join(upload_dir, file.filename)
            try:
                file.save(file_path)
                saved_file_size = os.path.getsize(file_path)
                file_info.append({'name': file.filename[:-4], 'size': saved_file_size})  # Remove .pdf extension
            except Exception as e:
                return jsonify({'success': False, 'message': f'Error saving file {file.filename}'})

        results, info = process_pdfs_in_folder(file_order, upload_dir, reg_numbers)
        df = pd.DataFrame(results).fillna('Not Found')
        df_transposed = df.transpose()
        components = df_transposed.columns.tolist()
        split_up_df = input_split_up(components, components_array)
        total_attainment = calculate_total_possible_attainment(split_up_df)

        split_up_df = calculate_row_wise_sum(split_up_df)
        split_up_df_without_sum = split_up_df.drop(columns=['Row-wise Sum'])

        attainment_table = calculate_attainment(split_up_df, df_transposed, reg_numbers)
        above_target_df = calculate_students_above_target(attainment_table, targetPercentage, total_attainment)
        print("\nAttainment Levels:")
        print(above_target_df)
        attainment_levels = get_attainment_levels(target_range)
        print("\nAttainment Levels:")
        print(attainment_levels)
        attainment_levels_df = assign_attainment_levels(above_target_df, attainment_levels)
        print("\nAttainment Levels Assigned:")
        # this values is not changing but slightly different
        print(attainment_levels_df)
        print("\nAttainment Levels Assigned:")
        df_clo_pso = pd.DataFrame(COPOMapperTablevalues)
        print(df_clo_pso)
        targets, df_attained = calculate_targets_and_attained(df_clo_pso, attainment_levels_df)
        print('hi')
        print(df_attained)
        print(attainment_table)
        print(targets)

        df_transposed = df_transposed.round(2)
        split_up_df_without_sum = split_up_df_without_sum.round(2)
        attainment_table = attainment_table.round(2)
        
        output_filename = 'results.xlsx'
        output_path = os.path.join(results_dir, output_filename)
        
        save_to_excel(df_transposed, info, total_attainment, attainment_table, split_up_df_without_sum, df_clo_pso, targets, df_attained, above_target_df,attainment_levels_df, output_path)
        
        response = {
            'success': True,
            'message': 'Attainments are calculated successfully.',
            'files': file_info,
            'download_url': f'/download/{output_filename}'
        }

        for file in uploaded_files:
            os.remove(os.path.join(upload_dir, file.filename))

        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)